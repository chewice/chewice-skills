#!/usr/bin/env python3
"""Append TSV rows to an xlsx sheet. Reads TSV from stdin.

Writes atomically (temp file + rename) to avoid corruption from
interrupted saves or sync conflicts on cloud-synced directories.
"""

import os
import sys
import tempfile
import shutil
from pathlib import Path

from openpyxl.styles import Font, PatternFill

KEEP_SHEETS = {"正常", "随访", "瑞美特"}


def _next_data_row(ws):
    """Find the next empty data row (col A is None), scanning upward."""
    for r in range(ws.max_row, 0, -1):
        if ws.cell(row=r, column=1).value is not None:
            return r + 1
    return 2


def _read_data_rows(ws):
    """Read all data rows (row >= 2) as list of lists."""
    if ws.max_row < 2:
        return []
    ncols = ws.max_column
    data = []
    for r in range(2, ws.max_row + 1):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, ncols + 1)]
        if any(v is not None for v in row_vals):
            data.append(row_vals)
    return data


def _delete_all_data_rows(ws):
    """Delete all rows from row 2 to max_row, then reset dimensions."""
    if ws.max_row < 2:
        return
    ws.delete_rows(2, ws.max_row - 1)


def _parse_table_date(val):
    """Parse 2026.8.27 / 2026/8/27 / 2026-8-27 into a sortable tuple."""
    if val is None or str(val).strip() == "":
        return (9999, 99, 99)
    text = str(val).strip()
    for sep in (".", "/", "-"):
        parts = text.split(sep)
        if len(parts) == 3:
            try:
                return (int(parts[0]), int(parts[1]), int(parts[2]))
            except (ValueError, TypeError):
                pass
    return (9999, 99, 99)


def _rewrite_data_rows(ws, data):
    """Write data rows starting at row 2."""
    for ri, row_vals in enumerate(data, 2):
        for ci, val in enumerate(row_vals, 1):
            if val is not None:
                ws.cell(row=ri, column=ci, value=val)


def _sort_followup_sheet(ws):
    """Sort follow-up sheet by col B (date), preserving V1 yellow fills."""
    if ws.max_row < 2:
        return
    data = _read_data_rows(ws)
    yellow_keys = set()
    for r, row_vals in enumerate(data, 2):
        fill = ws.cell(row=r, column=2).fill
        color = getattr(fill, "fgColor", None)
        rgb = getattr(color, "rgb", None) if color is not None else None
        if rgb and str(rgb).upper().endswith("FFFF00"):
            yellow_keys.add(tuple(row_vals))

    data.sort(key=lambda row_vals: _parse_table_date(row_vals[1] if len(row_vals) > 1 else None))
    _delete_all_data_rows(ws)
    _rewrite_data_rows(ws, data)
    for r, row_vals in enumerate(data, 2):
        if tuple(row_vals) in yellow_keys:
            ws.cell(row=r, column=2).fill = V1_HIGHLIGHT


def _sort_normal_sheet(ws):
    """Sort normal sheet by col C (assess date) then col A (id)."""
    if ws.max_row < 2:
        return
    data = _read_data_rows(ws)

    def _parse_id(val):
        if val is None or str(val).strip() == "":
            return float("inf")
        try:
            return float(val)
        except (ValueError, TypeError):
            return float("inf")

    def sort_key(row_vals):
        date_val = row_vals[2] if len(row_vals) > 2 else None
        id_val = row_vals[0] if row_vals else None
        return (_parse_table_date(date_val), _parse_id(id_val))

    data.sort(key=sort_key)
    _delete_all_data_rows(ws)
    _rewrite_data_rows(ws, data)


FONT_MAP = {
    "正常": Font(name="微软雅黑", size=11),
    "随访": Font(name="宋体", size=11),
    "瑞美特": Font(name="等线", size=11),
}

V1_HIGHLIGHT = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")


def _apply_v1_highlight(ws, flagged_rows):
    """Highlight date cell (col B) yellow for rows tagged with V1."""
    for r in flagged_rows:
        ws.cell(row=r, column=2).fill = V1_HIGHLIGHT


def _apply_font(ws, sheet_name):
    """Apply sheet-specific font to all cells."""
    font = FONT_MAP.get(sheet_name)
    if font is None:
        return
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            if cell.value is not None:
                cell.font = font


def append_tsv_to_xlsx(xlsx_path, sheet_name, tsv_text):
    import openpyxl

    lines = tsv_text.strip().split("\n")
    if not lines:
        return 0
    headers = [c.strip() for c in lines[0].split("\t")]
    rows = [[c.strip() for c in l.split("\t")] for l in lines[1:]]

    xlsx = Path(xlsx_path)
    if xlsx.exists():
        wb = openpyxl.load_workbook(str(xlsx))
        for sn in list(wb.sheetnames):
            if sn not in KEEP_SHEETS:
                del wb[sn]
    else:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(title=sheet_name)

    # Detect V1 tag column for followup sheets (not written to xlsx)
    v1_col_idx = None
    xlsx_headers = list(headers)
    if sheet_name == "随访" and "V1标记" in headers:
        v1_col_idx = headers.index("V1标记")
        xlsx_headers = [h for h in headers if h != "V1标记"]

    if sheet_name not in wb.sheetnames:
        for ci, h in enumerate(xlsx_headers, 1):
            ws.cell(row=1, column=ci, value=h)

    next_row = _next_data_row(ws)
    v1_flagged_rows = []

    for row in rows:
        xlsx_ci = 0
        for ci, val in enumerate(row, 1):
            if v1_col_idx is not None and ci == v1_col_idx + 1:
                if val == "是":
                    v1_flagged_rows.append(next_row)
                continue
            xlsx_ci += 1
            if val:
                ws.cell(row=next_row, column=xlsx_ci, value=val)
        next_row += 1

    if sheet_name == "正常" and ws.max_row > 1:
        _sort_normal_sheet(ws)

    _apply_font(ws, sheet_name)
    if sheet_name == "随访":
        _apply_v1_highlight(ws, v1_flagged_rows)
        if ws.max_row > 1:
            _sort_followup_sheet(ws)

    # Atomic write
    tmp_fd, tmp_path = tempfile.mkstemp(
        suffix=".xlsx", dir=str(xlsx.parent), prefix=".wip-"
    )
    os.close(tmp_fd)
    try:
        wb.save(tmp_path)
        shutil.move(tmp_path, str(xlsx))
    except BaseException:
        try:
            os.unlink(tmp_path)
        except OSError:
            pass
        raise

    return len(rows)


if __name__ == "__main__":
    import argparse

    p = argparse.ArgumentParser()
    p.add_argument("xlsx_path")
    p.add_argument("sheet_name")
    args = p.parse_args()
    tsv = sys.stdin.read()
    n = append_tsv_to_xlsx(args.xlsx_path, args.sheet_name, tsv)
    print(f"{args.sheet_name}: {n} rows")
